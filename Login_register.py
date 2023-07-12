import openpyxl

global wb 
wb = openpyxl.load_workbook("Records.xlsx")

global sh1 
sh1 = wb['Record1']

class Log_in:

    def __init__(self , email , password): 
        self.email = email
        self.password = password # we will take email and password as entry form the input widgets 

    def checker(self):        
        row = sh1.max_row
        for i in range( 2 , row + 1 ):
            if( (sh1.cell(i,1).value) == self.email ):
                if( (sh1.cell(i,2).value) == self.password ):
                    return i
                else :
                    return -1 
            
        if( i == row + 1 ):
            return 0
    # def new_pass(self,password):
        # info[self.email][0] = password

class Registration(Log_in):
    
    def __init__(self, email, password, name, phno, School):
        Log_in.__init__(self , email, password)
        self.name = name 
        self.phno = phno
        self.School = School

    def first_registration(self): # First we have taken input of name, password , email and all and then we are appending it to the dictionary
        sh1.cell( row = ( sh1.max_row ) + 1 , column = 1 , value = self.email )
        sh1.cell( row = ( sh1.max_row ) , column = 2 , value = self.password )
        sh1.cell( row = ( sh1.max_row ) , column = 3 , value = self.name )
        sh1.cell( row = ( sh1.max_row ) , column = 4 , value = self.phno )
        sh1.cell( row = ( sh1.max_row ) , column = 5 , value = self.School )
        sh1.cell( row = ( sh1.max_row ) , column = 9 , value = 0 )
        wb.save("Records.xlsx")
